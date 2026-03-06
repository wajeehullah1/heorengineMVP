"""AI-powered abstract screener for systematic literature reviews.

Loads the screening prompt template, formats it with PICO criteria and
abstract batches, calls the Claude API, parses structured responses into
:class:`~engines.slr.schema.ScreeningDecision` objects, and persists results.

Typical usage::

    from engines.slr.schema import Abstract, PICOCriteria
    from engines.slr.screener import create_screening_batch, screen_abstracts

    pico = PICOCriteria(
        population="Adults with type 2 diabetes",
        intervention="Remote CGM",
        comparison="Standard care",
        outcomes=["HbA1c reduction"],
        study_types=["RCT", "Economic evaluation"],
    )
    abstracts = [Abstract(pmid="12345678", ...)]

    batch   = create_screening_batch(abstracts, pico)
    decisions = screen_abstracts(abstracts, pico, batch_size=10)
    for d in decisions:
        batch.add_decision(d)

    filepath = export_screening_results(batch, format="csv")
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import anthropic
    _ANTHROPIC_AVAILABLE = True
except ImportError:  # pragma: no cover
    _ANTHROPIC_AVAILABLE = False
    anthropic = None  # type: ignore[assignment]

from engines.slr.schema import (
    Abstract,
    Confidence,
    Decision,
    PICOCriteria,
    PICOMatchItem,
    ScreeningBatch,
    ScreeningDecision,
)

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY: Optional[str] = os.getenv("ANTHROPIC_API_KEY")

# Model used for abstract screening.  Sonnet gives the best balance of
# reasoning quality and throughput for high-volume literature screening.
_MODEL = "claude-sonnet-4-6"

# Maximum tokens per API response.  A batch of 10 abstracts typically uses
# 1,500–2,500 tokens; 4,000 gives comfortable headroom.
_MAX_TOKENS = 4_000

# Retry configuration for transient API errors.
_MAX_RETRIES = 3
_RETRY_BASE_DELAY = 2.0  # seconds; doubled on each retry (exponential back-off)

# ── Paths ────────────────────────────────────────────────────────────────────

_REPO_ROOT   = Path(__file__).resolve().parent.parent.parent
_PROMPT_PATH = _REPO_ROOT / "agents" / "prompts" / "screening.txt"
_BATCHES_DIR = _REPO_ROOT / "data" / "slr" / "batches"
_EXPORTS_DIR = _REPO_ROOT / "data" / "slr" / "exports"

# ── Prompt section marker ────────────────────────────────────────────────────

# The template is split here to separate the static instructions from the
# abstract injection section, which is rebuilt for each batch.
_ABSTRACT_SECTION_MARKER = "ABSTRACT TO SCREEN"

# Default exclusion types used in {excluded_types} when not overridden.
_DEFAULT_EXCLUDED_TYPES = (
    "Conference abstract | Editorial | Letter | Commentary | "
    "Animal study | Case report (n < 10)"
)

# ── Compiled regex patterns for response parsing ─────────────────────────────

_DECISION_RE    = re.compile(r"Decision\s*:\s*(INCLUDE|EXCLUDE|UNCERTAIN)", re.IGNORECASE)
_CONFIDENCE_RE  = re.compile(r"Confidence\s*:\s*(HIGH|MEDIUM|LOW)",          re.IGNORECASE)
_PICO_LINE_RE   = re.compile(
    r"-\s*(Population|Intervention|Comparison|Outcome)\s+match\s*:"
    r"\s*(YES|NO|PARTIAL|N/A)\s*[—–\-]+\s*(.+)",
    re.IGNORECASE,
)
_REASONING_RE   = re.compile(
    r"Reasoning\s*:\s*(.+?)(?=\n\s*\n|\nExclusion|\Z)",
    re.DOTALL | re.IGNORECASE,
)
_EXCLUSION_RE   = re.compile(
    r"Exclusion reasons.*?:\s*\n((?:\s*-\s*.+\n?)+)",
    re.IGNORECASE | re.DOTALL,
)


# ═══════════════════════════════════════════════════════════════════════════════
# 1. Prompt loading
# ═══════════════════════════════════════════════════════════════════════════════

def load_screening_prompt() -> str:
    """Load the raw screening prompt template from ``agents/prompts/screening.txt``.

    Returns:
        The complete template string with ``{placeholder}`` variables
        ready for :meth:`str.format` substitution.

    Raises:
        FileNotFoundError: If the prompt file does not exist at the expected path.
    """
    if not _PROMPT_PATH.is_file():
        raise FileNotFoundError(
            f"Screening prompt not found at {_PROMPT_PATH}. "
            "Ensure agents/prompts/screening.txt exists relative to the repo root."
        )
    template = _PROMPT_PATH.read_text(encoding="utf-8")
    logger.debug("Loaded screening prompt (%d chars) from %s", len(template), _PROMPT_PATH)
    return template


# ═══════════════════════════════════════════════════════════════════════════════
# 2. Prompt formatting
# ═══════════════════════════════════════════════════════════════════════════════

def format_screening_prompt(
    pico: PICOCriteria,
    abstracts: list[Abstract],
) -> str:
    """Build a complete screening prompt ready to send to the Claude API.

    The function:

    1. Loads the template via :func:`load_screening_prompt`.
    2. Fills in all PICO-related ``{placeholders}`` in the instructions
       section (everything before the ``ABSTRACT TO SCREEN`` marker).
    3. Replaces the single-abstract section with a numbered multi-abstract
       block and updated output instructions for batched responses.

    Args:
        pico:      PICO eligibility criteria for this screening run.
        abstracts: One or more abstracts to include in the prompt.  When
                   called from :func:`screen_abstracts` these will be one
                   batch slice, not the full corpus.

    Returns:
        A formatted string ready to use as the ``content`` of a Claude
        user message.

    Raises:
        FileNotFoundError: Propagated from :func:`load_screening_prompt`.
        ValueError:        If *abstracts* is empty.
    """
    if not abstracts:
        raise ValueError("abstracts must contain at least one Abstract")

    template = load_screening_prompt()

    # ── Split at the abstract section marker ────────────────────────────────
    if _ABSTRACT_SECTION_MARKER in template:
        instructions_raw = template.split(_ABSTRACT_SECTION_MARKER)[0]
    else:
        logger.warning(
            "Abstract section marker %r not found in template; using full template",
            _ABSTRACT_SECTION_MARKER,
        )
        instructions_raw = template

    # ── Extra exclusions from PICO.exclusion_criteria ──────────────────────
    extra_exclusions = "\n".join(
        f"  - {e}" for e in (pico.exclusion_criteria or [])
    )

    # ── Fill PICO placeholders ──────────────────────────────────────────────
    # Abstract-specific keys ({pmid}, {title}, etc.) also appear in the
    # comment header of the template.  Supply generic text so str.format()
    # doesn't raise KeyError.
    instructions = instructions_raw.format(
        population=pico.population,
        population_details=pico.population,
        intervention=pico.intervention,
        intervention_details=pico.intervention,
        comparison=pico.comparison,
        comparison_details=pico.comparison,
        outcomes=", ".join(pico.outcomes),
        outcome_list=" | ".join(pico.outcomes),
        study_types=", ".join(pico.study_types),
        acceptable_types=" | ".join(pico.study_types),
        excluded_types=_DEFAULT_EXCLUDED_TYPES,
        additional_exclusions=extra_exclusions,
        # Abstract-specific keys (appear only in comment header in this section)
        pmid="<varies per abstract — see below>",
        title="<varies per abstract — see below>",
        authors="<varies per abstract — see below>",
        journal="<varies per abstract — see below>",
        abstract_text="<varies per abstract — see below>",
    )

    # ── Build numbered abstract block ───────────────────────────────────────
    n = len(abstracts)
    abstract_lines: list[str] = [
        "=" * 80,
        f"ABSTRACTS TO SCREEN  ({n} abstract{'s' if n != 1 else ''})",
        "=" * 80,
        "",
        "Screen each abstract below in order. Provide one complete decision",
        "block per abstract, clearly prefixed with its PMID as shown.",
        "",
    ]

    for idx, ab in enumerate(abstracts, 1):
        abstract_lines += [
            f"── Abstract {idx} of {n} " + "─" * max(0, 60 - len(str(idx)) - len(str(n))),
            f"PMID:     {ab.pmid}",
            f"Title:    {ab.title}",
            f"Authors:  {ab.short_citation()}",
            f"Source:   {ab.journal} ({ab.year})",
            "",
            "Abstract:",
            ab.abstract,
            "",
        ]

    # ── Output instructions for batched response ────────────────────────────
    abstract_lines += [
        "=" * 80,
        "END OF PROMPT — Provide your structured response below.",
        "",
        "For EACH abstract above output one block in this exact format",
        "(repeat the block for every abstract, separated by a blank line):",
        "",
        "PMID: [pmid]",
        "Decision: [INCLUDE | EXCLUDE | UNCERTAIN]",
        "Confidence: [HIGH | MEDIUM | LOW]",
        "",
        "PICO Assessment:",
        "- Population match:    [YES | NO | PARTIAL] — [brief note]",
        "- Intervention match:  [YES | NO | PARTIAL] — [brief note]",
        "- Comparison match:    [YES | NO | PARTIAL | N/A] — [brief note]",
        "- Outcome match:       [YES | NO | PARTIAL] — [brief note]",
        "",
        "Reasoning: [2–3 sentences referencing specific abstract content]",
        "",
        "Exclusion reasons (if EXCLUDE, omit otherwise):",
        "- [specific reason]",
        "=" * 80,
    ]

    return instructions + "\n".join(abstract_lines)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. Response parsing
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_pmid_block(response_text: str, pmid: str) -> str:
    """Return the slice of *response_text* that corresponds to *pmid*.

    Claude's batch response contains one decision block per abstract,
    each prefixed with ``PMID: <value>``.  This function splits on those
    headers and returns the block whose PMID matches the requested one.

    Falls back to the full response text if no PMID header is found (e.g.
    single-abstract responses without an explicit PMID prefix).

    Args:
        response_text: Full text returned by the Claude API for a batch.
        pmid:          The PubMed ID whose block is to be extracted.

    Returns:
        Substring of *response_text* containing the relevant decision block.
    """
    # Split into per-abstract blocks on any "PMID:" header
    blocks = re.split(r"(?m)^(?=PMID\s*:)", response_text)

    # Filter to non-empty blocks only
    pmid_blocks = [b for b in blocks if b.strip()]

    for block in pmid_blocks:
        if re.match(rf"PMID\s*:\s*{re.escape(pmid)}\b", block.strip(), re.IGNORECASE):
            return block

    # If there are no PMID headers at all (single-abstract response without
    # the PMID prefix), fall back to the full text so parsing still works.
    if not pmid_blocks or all(
        not re.match(r"PMID\s*:", b.strip(), re.IGNORECASE) for b in pmid_blocks
    ):
        logger.debug("No PMID headers found; treating full response as single-abstract block")
        return response_text

    # Multiple blocks exist but none matched — PMID is genuinely absent.
    # Return empty string so parse_screening_response produces UNCERTAIN/LOW.
    logger.warning(
        "PMID %r not found among %d response block(s); will return UNCERTAIN",
        pmid, len(pmid_blocks),
    )
    return ""


def parse_screening_response(response_text: str, pmid: str) -> ScreeningDecision:
    """Parse Claude's structured text response for a single abstract.

    Extracts the decision block for *pmid* from *response_text* (which may
    contain decisions for multiple abstracts) then parses:

    - ``Decision`` → :class:`~engines.slr.schema.Decision` enum
    - ``Confidence`` → :class:`~engines.slr.schema.Confidence` enum
    - ``PICO Assessment`` lines → ``dict[str, PICOMatchItem]``
    - ``Reasoning`` → free-text explanation
    - ``Exclusion reasons`` → list of strings (excluded decisions only)

    Any PICO component that is absent from the response receives a default
    ``PICOMatchItem(matched=False, note="Not assessed in response")``.

    When the response is too malformed to extract a meaningful decision,
    returns ``Decision.UNCERTAIN`` with ``Confidence.LOW`` so the abstract
    proceeds to human review rather than being silently dropped.

    Args:
        response_text: Raw text from the Claude API (may cover multiple abstracts).
        pmid:          PubMed ID of the abstract being parsed.

    Returns:
        A fully-populated :class:`~engines.slr.schema.ScreeningDecision`.
    """
    block = _extract_pmid_block(response_text, pmid)

    # ── Decision ────────────────────────────────────────────────────────────
    decision_match = _DECISION_RE.search(block)
    if decision_match:
        decision = Decision(decision_match.group(1).lower())
    else:
        logger.warning("No Decision field found for PMID %s; defaulting to UNCERTAIN", pmid)
        decision = Decision.UNCERTAIN

    # ── Confidence ──────────────────────────────────────────────────────────
    confidence_match = _CONFIDENCE_RE.search(block)
    if confidence_match:
        confidence = Confidence(confidence_match.group(1).lower())
    else:
        confidence = Confidence.LOW

    # ── PICO assessment ─────────────────────────────────────────────────────
    pico_match: dict[str, PICOMatchItem] = {}
    _component_map = {
        "population":   "population",
        "intervention": "intervention",
        "comparison":   "comparison",
        "outcome":      "outcome",
    }
    for m in _PICO_LINE_RE.finditer(block):
        component = m.group(1).lower()
        verdict   = m.group(2).upper()
        note      = m.group(3).strip()
        if component in _component_map:
            pico_match[_component_map[component]] = PICOMatchItem(
                # YES or PARTIAL → consider matched for scoring purposes
                matched=verdict in ("YES", "PARTIAL"),
                note=note[:300],  # cap to prevent runaway notes
            )

    # Fill any missing PICO components with a safe default
    for key in ("population", "intervention", "comparison", "outcome"):
        if key not in pico_match:
            pico_match[key] = PICOMatchItem(
                matched=False,
                note="Component absent from screener response — verify manually",
            )

    # ── Reasoning ───────────────────────────────────────────────────────────
    reasoning_match = _REASONING_RE.search(block)
    if reasoning_match:
        reasoning = reasoning_match.group(1).strip()
    else:
        reasoning = (
            "No structured reasoning provided. "
            "Review the raw response for this abstract manually."
        )

    # ── Exclusion reasons ───────────────────────────────────────────────────
    exclusion_reasons: list[str] = []
    excl_match = _EXCLUSION_RE.search(block)
    if excl_match:
        for line in excl_match.group(1).splitlines():
            reason = re.sub(r"^\s*-\s*", "", line).strip()
            if reason:
                exclusion_reasons.append(reason)

    # ── Assemble ScreeningDecision ──────────────────────────────────────────
    try:
        return ScreeningDecision(
            pmid=pmid,
            decision=decision,
            confidence=confidence,
            reasoning=reasoning,
            pico_match=pico_match,
            exclusion_reasons=exclusion_reasons,
        )
    except Exception as exc:
        logger.error(
            "Failed to construct ScreeningDecision for PMID %s: %s. "
            "Returning UNCERTAIN/LOW fallback.",
            pmid, exc,
        )
        return ScreeningDecision(
            pmid=pmid,
            decision=Decision.UNCERTAIN,
            confidence=Confidence.LOW,
            reasoning=(
                f"Parse error — manual review required. "
                f"Original error: {exc}. "
                f"Raw block (first 300 chars): {block[:300]!r}"
            ),
            pico_match={
                k: PICOMatchItem(matched=False, note="Parse error — see reasoning")
                for k in ("population", "intervention", "comparison", "outcome")
            },
            exclusion_reasons=[],
        )


# ═══════════════════════════════════════════════════════════════════════════════
# 4. API call with retry
# ═══════════════════════════════════════════════════════════════════════════════

def _call_claude_with_retry(client: "anthropic.Anthropic", prompt: str) -> str:  # type: ignore[name-defined]
    """Send *prompt* to Claude and return the response text, with retries.

    Retries on:

    - ``RateLimitError`` — exponential back-off up to :data:`_MAX_RETRIES` attempts.
    - ``APIConnectionError`` — same back-off; catches transient network failures.
    - ``APIStatusError`` with HTTP 5xx — server-side errors that may be transient.

    Non-retriable errors (HTTP 4xx excluding 429) are re-raised immediately.

    Args:
        client: An authenticated :class:`anthropic.Anthropic` client instance.
        prompt: The fully-formatted screening prompt to send as a user message.

    Returns:
        The raw text content of Claude's first response block.

    Raises:
        anthropic.APIError: If all retries are exhausted or the error is
                            not retriable.
    """
    last_exc: Optional[Exception] = None

    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            message = client.messages.create(
                model=_MODEL,
                max_tokens=_MAX_TOKENS,
                messages=[{"role": "user", "content": prompt}],
            )
            return message.content[0].text

        except anthropic.RateLimitError as exc:  # type: ignore[attr-defined]
            delay = _RETRY_BASE_DELAY * (2 ** (attempt - 1))
            logger.warning(
                "Rate limited (attempt %d/%d) — retrying in %.1f s",
                attempt, _MAX_RETRIES, delay,
            )
            last_exc = exc
            time.sleep(delay)

        except anthropic.APIConnectionError as exc:  # type: ignore[attr-defined]
            delay = _RETRY_BASE_DELAY * (2 ** (attempt - 1))
            logger.warning(
                "Connection error (attempt %d/%d) — retrying in %.1f s: %s",
                attempt, _MAX_RETRIES, delay, exc,
            )
            last_exc = exc
            time.sleep(delay)

        except anthropic.APIStatusError as exc:  # type: ignore[attr-defined]
            if exc.status_code >= 500:
                delay = _RETRY_BASE_DELAY * (2 ** (attempt - 1))
                logger.warning(
                    "Server error HTTP %d (attempt %d/%d) — retrying in %.1f s",
                    exc.status_code, attempt, _MAX_RETRIES, delay,
                )
                last_exc = exc
                time.sleep(delay)
            else:
                # 4xx (except 429 handled above) are programming or auth errors
                raise

    raise last_exc  # type: ignore[misc]


# ═══════════════════════════════════════════════════════════════════════════════
# 5. Main screening function
# ═══════════════════════════════════════════════════════════════════════════════

def screen_abstracts(
    abstracts: list[Abstract],
    pico: PICOCriteria,
    batch_size: int = 10,
) -> list[ScreeningDecision]:
    """Screen a list of abstracts against PICO criteria using the Claude API.

    Splits *abstracts* into batches of up to *batch_size*, sends each batch
    as a single API call, then parses the structured response into
    :class:`~engines.slr.schema.ScreeningDecision` objects.

    If parsing fails for an individual abstract it is returned as
    ``Decision.UNCERTAIN`` / ``Confidence.LOW`` so it proceeds to human
    review rather than being lost.

    Args:
        abstracts:  All abstracts to screen.  May be any length.
        pico:       PICO criteria used to evaluate eligibility.
        batch_size: Number of abstracts per API call.  Reduce if responses
                    are being truncated (lower ``max_tokens`` headroom).
                    Increase to reduce API call count for large corpora.

    Returns:
        One :class:`~engines.slr.schema.ScreeningDecision` per abstract,
        in the same order as *abstracts*.

    Raises:
        ImportError:    If the ``anthropic`` package is not installed.
        EnvironmentError: If ``ANTHROPIC_API_KEY`` is not set.
        anthropic.APIError: For non-retriable API failures.
    """
    if not _ANTHROPIC_AVAILABLE:
        raise ImportError(
            "The 'anthropic' package is required for abstract screening. "
            "Install it with: pip install anthropic"
        )
    # Read the key lazily so it picks up values exported after module import
    # (e.g. via `source .env` before running pytest).
    api_key = os.getenv("ANTHROPIC_API_KEY") or ANTHROPIC_API_KEY
    if not api_key:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY environment variable is not set. "
            "Export it before calling screen_abstracts()."
        )

    client = anthropic.Anthropic(api_key=api_key)  # type: ignore[union-attr]

    all_decisions: list[ScreeningDecision] = []
    batches = [
        abstracts[i : i + batch_size]
        for i in range(0, len(abstracts), batch_size)
    ]

    logger.info(
        "Screening %d abstract(s) in %d batch(es) of up to %d (model: %s)",
        len(abstracts), len(batches), batch_size, _MODEL,
    )

    for batch_num, batch in enumerate(batches, 1):
        logger.info(
            "Sending batch %d/%d (%d abstracts) to Claude API",
            batch_num, len(batches), len(batch),
        )
        prompt = format_screening_prompt(pico, batch)

        try:
            response_text = _call_claude_with_retry(client, prompt)
        except Exception as exc:
            logger.error(
                "Batch %d/%d failed after %d retries: %s — "
                "returning UNCERTAIN/LOW for all %d abstracts in this batch",
                batch_num, len(batches), _MAX_RETRIES, exc, len(batch),
            )
            for ab in batch:
                all_decisions.append(
                    ScreeningDecision(
                        pmid=ab.pmid,
                        decision=Decision.UNCERTAIN,
                        confidence=Confidence.LOW,
                        reasoning=(
                            f"API call failed — manual review required. "
                            f"Error: {exc}"
                        ),
                        pico_match={
                            k: PICOMatchItem(matched=False, note="API failure")
                            for k in ("population", "intervention", "comparison", "outcome")
                        },
                    )
                )
            continue

        logger.debug("Received %d-char response for batch %d", len(response_text), batch_num)

        for ab in batch:
            decision = parse_screening_response(response_text, ab.pmid)
            all_decisions.append(decision)
            logger.debug(
                "PMID %s → %s (%s, PICO score %d/4)",
                ab.pmid, decision.decision.value,
                decision.confidence.value, decision.pico_match_score,
            )

    included  = sum(1 for d in all_decisions if d.decision == Decision.INCLUDE)
    excluded  = sum(1 for d in all_decisions if d.decision == Decision.EXCLUDE)
    uncertain = sum(1 for d in all_decisions if d.decision == Decision.UNCERTAIN)
    logger.info(
        "Screening complete: %d total — %d included, %d excluded, %d uncertain",
        len(all_decisions), included, excluded, uncertain,
    )

    return all_decisions


# ═══════════════════════════════════════════════════════════════════════════════
# 6. Batch persistence
# ═══════════════════════════════════════════════════════════════════════════════

def create_screening_batch(
    abstracts: list[Abstract],
    pico: PICOCriteria,
) -> ScreeningBatch:
    """Create a :class:`~engines.slr.schema.ScreeningBatch` and persist it to disk.

    Generates a timestamp-prefixed UUID batch ID of the form::

        slr_20240315_143022_a1b2c3d4

    Saves the batch (without decisions, which are appended later) as a JSON
    file under ``data/slr/batches/``.

    Args:
        abstracts: Abstracts to include in the batch.
        pico:      PICO criteria for this screening run.

    Returns:
        A new :class:`~engines.slr.schema.ScreeningBatch` with an empty
        decisions list and ``summary`` initialised to all zeros.

    Raises:
        OSError: If the batch file cannot be written.
    """
    ts = datetime.now(timezone.utc)
    batch_id = f"slr_{ts.strftime('%Y%m%d_%H%M%S')}_{str(uuid.uuid4())[:8]}"

    batch = ScreeningBatch(
        batch_id=batch_id,
        pico_criteria=pico,
        abstracts=abstracts,
    )

    _BATCHES_DIR.mkdir(parents=True, exist_ok=True)
    batch_path = _BATCHES_DIR / f"{batch_id}.json"
    batch_path.write_text(
        batch.model_dump_json(indent=2),
        encoding="utf-8",
    )
    logger.info(
        "Saved batch %s (%d abstracts) to %s",
        batch_id, len(abstracts), batch_path,
    )
    return batch


def save_batch(batch: ScreeningBatch) -> Path:
    """Overwrite the persisted JSON for *batch* with its current state.

    Call this after appending decisions so the file on disk stays in sync
    with the in-memory object.

    Args:
        batch: The batch to persist (must have a valid :attr:`~ScreeningBatch.batch_id`).

    Returns:
        The :class:`~pathlib.Path` of the written file.

    Raises:
        OSError: If the file cannot be written.
    """
    _BATCHES_DIR.mkdir(parents=True, exist_ok=True)
    batch_path = _BATCHES_DIR / f"{batch.batch_id}.json"
    batch_path.write_text(
        batch.model_dump_json(indent=2),
        encoding="utf-8",
    )
    logger.info("Updated batch %s on disk (%s decisions)", batch.batch_id, len(batch.decisions))
    return batch_path


def load_batch(batch_id: str) -> ScreeningBatch:
    """Load a previously saved :class:`~engines.slr.schema.ScreeningBatch` from disk.

    Args:
        batch_id: The batch identifier, e.g. ``'slr_20240315_143022_a1b2c3d4'``.

    Returns:
        The deserialised :class:`~engines.slr.schema.ScreeningBatch`.

    Raises:
        FileNotFoundError: If no batch file exists for *batch_id*.
        pydantic.ValidationError: If the JSON does not match the schema.
    """
    batch_path = _BATCHES_DIR / f"{batch_id}.json"
    if not batch_path.is_file():
        raise FileNotFoundError(
            f"No batch file found for batch_id={batch_id!r}. "
            f"Expected: {batch_path}"
        )
    data = json.loads(batch_path.read_text(encoding="utf-8"))
    return ScreeningBatch.model_validate(data)


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Export
# ═══════════════════════════════════════════════════════════════════════════════

def export_screening_results(batch: ScreeningBatch, format: str = "csv") -> str:
    """Export screening decisions to CSV or Excel.

    Columns:

    ==================  ====================================================
    PMID                PubMed ID
    Title               Article title (from :attr:`Abstract.title`)
    Authors             Short citation string (e.g. ``'Smith JA et al. (2023)'``)
    Journal             Journal name
    Year                Publication year
    Decision            include / exclude / uncertain
    Confidence          high / medium / low
    PICO_Score          0–4 (matched PICO components)
    Reasoning           Screener's narrative explanation
    Exclusion_Reasons   Semicolon-separated exclusion reasons (if excluded)
    Reviewer            Reviewer identifier (default: ``'AI-Claude'``)
    Timestamp           ISO-8601 UTC timestamp of the decision
    ==================  ====================================================

    Files are written to ``data/slr/exports/``.

    Args:
        batch:  A :class:`~engines.slr.schema.ScreeningBatch` with at least
                one decision.
        format: ``'csv'`` (default) or ``'excel'`` / ``'xlsx'``.

    Returns:
        Absolute path to the written file as a string.

    Raises:
        ValueError:     If *format* is not ``'csv'``, ``'excel'``, or ``'xlsx'``.
        ImportError:    If ``openpyxl`` is not installed when format is Excel.
        OSError:        If the export file cannot be written.
    """
    _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Build abstract lookup for metadata columns
    abstract_by_pmid: dict[str, Abstract] = {a.pmid: a for a in batch.abstracts}

    _COLUMNS = [
        "PMID", "Title", "Authors", "Journal", "Year",
        "Decision", "Confidence", "PICO_Score",
        "Reasoning", "Exclusion_Reasons", "Reviewer", "Timestamp",
    ]

    rows: list[dict] = []
    for dec in batch.decisions:
        ab = abstract_by_pmid.get(dec.pmid)
        rows.append({
            "PMID":              dec.pmid,
            "Title":             ab.title             if ab else "",
            "Authors":           ab.short_citation()  if ab else "",
            "Journal":           ab.journal           if ab else "",
            "Year":              ab.year              if ab else "",
            "Decision":          dec.decision.value,
            "Confidence":        dec.confidence.value,
            "PICO_Score":        dec.pico_match_score,
            "Reasoning":         dec.reasoning,
            "Exclusion_Reasons": "; ".join(dec.exclusion_reasons),
            "Reviewer":          dec.reviewer,
            "Timestamp":         dec.timestamp.isoformat(),
        })

    fmt = format.strip().lower()

    # ── CSV ──────────────────────────────────────────────────────────────────
    if fmt == "csv":
        filepath = _EXPORTS_DIR / f"{batch.batch_id}.csv"
        with filepath.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)
        logger.info("Exported %d decisions to %s", len(rows), filepath)
        return str(filepath)

    # ── Excel ────────────────────────────────────────────────────────────────
    if fmt in ("excel", "xlsx"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        filepath = _EXPORTS_DIR / f"{batch.batch_id}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Screening Results"

        # Header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D2044", end_color="0D2044", fill_type="solid")
        for col_idx, col_name in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows with colour-coded Decision column
        _DECISION_COLOURS = {
            "include":  "D4EDDA",  # green tint
            "exclude":  "F8D7DA",  # red tint
            "uncertain": "FFF3CD",  # amber tint
        }
        decision_col_idx = _COLUMNS.index("Decision") + 1

        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(_COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            # Colour-code the Decision cell
            decision_val = row["Decision"]
            colour = _DECISION_COLOURS.get(decision_val, "FFFFFF")
            ws.cell(row=row_idx, column=decision_col_idx).fill = PatternFill(
                start_color=colour, end_color=colour, fill_type="solid"
            )

        # Auto-width approximation
        col_widths = {col: len(col) + 2 for col in _COLUMNS}
        for row in rows:
            for col in _COLUMNS:
                col_widths[col] = max(col_widths[col], min(len(str(row[col])), 60))
        for col_idx, col_name in enumerate(_COLUMNS, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_widths[col_name]

        wb.save(filepath)
        logger.info("Exported %d decisions to %s", len(rows), filepath)
        return str(filepath)

    raise ValueError(
        f"Unsupported export format: {format!r}. "
        "Use 'csv', 'excel', or 'xlsx'."
    )
